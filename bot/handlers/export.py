"""
Экспорт заказов в Excel-файл.
"""
from __future__ import annotations

import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def build_orders_excel(orders: list, items_by_order: dict) -> bytes:
    wb = openpyxl.Workbook()

    ws_orders = wb.active
    ws_orders.title = "Заказы"
    header = ["#", "Покупатель", "User ID", "Метод оплаты", "Stars", "Руб", "Статус", "Дата"]
    _write_header(ws_orders, header)

    for i, o in enumerate(orders, start=2):
        buyer = o["full_name"] or f"id{o['user_id']}"
        ws_orders.cell(i, 1, o["id"])
        ws_orders.cell(i, 2, buyer)
        ws_orders.cell(i, 3, o["user_id"])
        ws_orders.cell(i, 4, o["pay_method"])
        ws_orders.cell(i, 5, o["total_stars"])
        ws_orders.cell(i, 6, o["total_rub"])
        ws_orders.cell(i, 7, "✅ Оплачен" if o["status"] == "paid" else "⏳ Ожидает")
        ws_orders.cell(i, 8, o["created_at"])

    _auto_width(ws_orders)

    ws_items = wb.create_sheet("Состав")
    header2 = ["Заказ #", "Товар", "Кол-во", "Stars/шт", "Руб/шт", "Итого Stars", "Итого Руб"]
    _write_header(ws_items, header2)

    row = 2
    for order_id, items in items_by_order.items():
        for item in items:
            ws_items.cell(row, 1, order_id)
            ws_items.cell(row, 2, item["name"])
            ws_items.cell(row, 3, item["qty"])
            ws_items.cell(row, 4, item["price_stars"])
            ws_items.cell(row, 5, item["price_rub"])
            ws_items.cell(row, 6, item["price_stars"] * item["qty"])
            ws_items.cell(row, 7, item["price_rub"] * item["qty"])
            row += 1

    _auto_width(ws_items)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_header(ws, columns: list[str]) -> None:
    fill = PatternFill("solid", fgColor="D9D9D9")
    bold = Font(bold=True)
    for col, title in enumerate(columns, start=1):
        cell = ws.cell(1, col, title)
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
